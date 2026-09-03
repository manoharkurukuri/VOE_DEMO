"""Unit tests for ExcelService (app.service.excel_service)."""

import io

import pytest
from openpyxl import load_workbook

from app.core.constants import EXCEL_HEADERS
from app.core.exceptions import ExcelGenerationError
from app.service.excel_service import ExcelService


@pytest.fixture
def service(tmp_path):
    return ExcelService(storage_dir=tmp_path, timezone_name="Asia/Kolkata")


def test_slugify():
    assert ExcelService._slugify("Norm Reeves Honda!") == "norm_reeves_honda"
    assert ExcelService._slugify("   ") == "dealer"
    assert ExcelService._slugify("A/B  C") == "a_b_c"


def test_build_file_name_with_stem(service):
    assert service.build_file_name("Any", file_stem="TEST001_Subaru_20240101") == (
        "test001_subaru_20240101.xlsx"
    )


def test_build_file_name_from_dealer_and_date(service):
    name = service.build_file_name("Flow Subaru")
    assert name.startswith("flow_subaru_")
    assert name.endswith(".xlsx")


def test_build_workbook_bytes_produces_valid_xlsx(service, sample_incentive):
    file_name, data = service.build_workbook_bytes(
        dealer_name="Flow Subaru",
        incentives=[sample_incentive],
        source_url="https://example.com/specials",
        file_stem="flow_subaru_test",
    )
    assert file_name == "flow_subaru_test.xlsx"
    assert isinstance(data, bytes) and data

    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws.title == "Monthly Vehicle Incentives"
    # Header row matches the configured headers.
    header_row = [c.value for c in ws[1]]
    assert header_row == EXCEL_HEADERS
    # Exactly one data row was written.
    assert ws.max_row == 2
    # Source URL is attached as a comment on A1.
    assert ws["A1"].comment is not None
    assert "example.com/specials" in ws["A1"].comment.text


def test_finance_rate_is_converted_to_fraction(service, sample_incentive):
    # Column 19 (S) is Finance Rate; the service divides the percent by 100.
    _, data = service.build_workbook_bytes(
        dealer_name="Flow Subaru",
        incentives=[sample_incentive],
        file_stem="rate_test",
    )
    ws = load_workbook(io.BytesIO(data)).active
    assert ws.cell(row=2, column=19).value == pytest.approx(0.019)


def test_empty_incentives_writes_header_only(service):
    _, data = service.build_workbook_bytes(
        dealer_name="Empty Dealer",
        incentives=[],
        file_stem="empty",
    )
    ws = load_workbook(io.BytesIO(data)).active
    assert ws.max_row == 1
    assert [c.value for c in ws[1]] == EXCEL_HEADERS


def test_build_workbook_wraps_errors(monkeypatch, service, sample_incentive):
    # Force an internal failure and confirm it surfaces as ExcelGenerationError.
    import app.service.excel_service as mod

    class Boom:
        def __init__(self, *a, **k):
            raise RuntimeError("nope")

    monkeypatch.setattr(mod, "Workbook", Boom)
    with pytest.raises(ExcelGenerationError):
        service.build_workbook_bytes(
            dealer_name="X", incentives=[sample_incentive], file_stem="x"
        )
